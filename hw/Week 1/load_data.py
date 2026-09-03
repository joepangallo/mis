#!/usr/bin/env python3
"""
load_data.py - import a spreadsheet or text data file into a SQLite database.

This is the tool that replaces "File > Get External Data > Import" in Access.
You point it at a data file, tell it what to name the table, and it creates
that table inside a .db file and loads every row.

USAGE
-----
    python load_data.py <source file> <table name> <database file> [--sheet NAME]

EXAMPLES
--------
    python load_data.py data/FrequentFliers.txt frequent_fliers campus_travel.db
    python load_data.py data/TicketSales.csv    ticket_sales    campus_travel.db
    python load_data.py data/frequentflier2.xlsx customers      mileage.db --sheet Customers

WHAT IT DOES
------------
1. Reads the file. Tab-delimited .txt, comma-delimited .csv, and Excel .xlsx
   are all understood. The delimiter is detected, not guessed at by you.
2. Turns the header row into safe column names: "Frequent_Flier_Number"
   and "Meal Category" both become frequent_flier_number and meal_category.
3. Looks at the values in each column and picks a SQLite type - INTEGER,
   REAL, or TEXT. This is the same decision Access makes for you in its
   import wizard, except here you can see it happen and print it out.
4. Creates the table (dropping any earlier version of it) and inserts
   every row inside a single transaction.
5. Prints the schema it built and the row count, so you can confirm the
   import worked before you write a single query.

Only openpyxl is needed, and only if you are importing an .xlsx file:
    pip install openpyxl
"""

import argparse
import csv
import re
import sqlite3
import sys
from pathlib import Path


# --------------------------------------------------------------------------
# Step 1: reading the source file
# --------------------------------------------------------------------------

def read_delimited(path):
    """Read a .csv or .txt file and return (header, rows).

    csv.Sniffer inspects the first chunk of the file and works out whether
    the columns are separated by commas or tabs, so the same function
    handles TicketSales.csv and FrequentFliers.txt without being told which
    is which.
    """
    with open(path, newline="", encoding="utf-8-sig") as f:
        sample = f.read(8192)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        except csv.Error:
            dialect = csv.excel_tab if "\t" in sample else csv.excel
        rows = [r for r in csv.reader(f, dialect)]

    rows = [r for r in rows if any(str(cell).strip() for cell in r)]
    if not rows:
        sys.exit(f"'{path}' has no data in it.")
    return rows[0], rows[1:]


def read_excel(path, sheet_name=None):
    """Read one worksheet of an .xlsx file and return (header, rows)."""
    try:
        import openpyxl
    except ImportError:
        sys.exit("Reading .xlsx files needs openpyxl. Run:  pip install openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name is None:
        ws = wb[wb.sheetnames[0]]
    elif sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        sys.exit(f"No sheet named '{sheet_name}'. This file has: {', '.join(wb.sheetnames)}")

    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    rows = [r for r in rows if any(c is not None and str(c).strip() for c in r)]
    if not rows:
        sys.exit(f"Sheet '{ws.title}' has no data in it.")
    return rows[0], rows[1:]


# --------------------------------------------------------------------------
# Step 2: turning headers into legal column names
# --------------------------------------------------------------------------

def clean_column_name(raw, position):
    """'Meal Category' -> meal_category,  'Frequent_Flier_Number' -> frequent_flier_number.

    Spaces and punctuation become underscores, everything goes to lower
    case, and a name that would start with a digit gets a prefix. A blank
    header becomes column_1, column_2, and so on, so nothing is ever
    silently dropped.
    """
    name = str(raw).strip().lower() if raw is not None else ""
    name = re.sub(r"[^a-z0-9]+", "_", name).strip("_")
    if not name:
        name = f"column_{position}"
    if name[0].isdigit():
        name = f"c_{name}"
    return name


def unique_names(headers):
    """Make sure no two columns end up with the same name."""
    out, seen = [], {}
    for i, h in enumerate(headers, start=1):
        name = clean_column_name(h, i)
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 1
        out.append(name)
    return out


# --------------------------------------------------------------------------
# Step 3: choosing a column type
# --------------------------------------------------------------------------

def looks_like_integer(v):
    return bool(re.fullmatch(r"-?\d+", v))


def looks_like_decimal(v):
    return bool(re.fullmatch(r"-?\d*\.\d+", v))


def infer_type(values):
    """Decide whether a column is INTEGER, REAL, or TEXT.

    A column is only numeric if EVERY non-blank value in it is numeric.
    That rule matters here: a phone number like 509-332-4578 contains
    digits but is not a number, and a frequent flier number is an
    identifier you will never do arithmetic on. Both correctly land in
    TEXT, which is what you want - leading zeros survive and nothing gets
    rounded.
    """
    seen = [str(v).strip() for v in values if v is not None and str(v).strip() != ""]
    if not seen:
        return "TEXT"
    if all(looks_like_integer(v) for v in seen):
        return "INTEGER"
    if all(looks_like_integer(v) or looks_like_decimal(v) for v in seen):
        return "REAL"
    return "TEXT"


def convert(value, sql_type):
    """Turn one spreadsheet cell into the Python value SQLite should store."""
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    if sql_type == "INTEGER":
        return int(float(text))
    if sql_type == "REAL":
        return float(text)
    return text


# --------------------------------------------------------------------------
# Step 4: building the table and loading it
# --------------------------------------------------------------------------

def load(source, table, database, sheet=None, quiet=False, text_columns=()):
    source = Path(source)
    if not source.exists():
        sys.exit(f"Cannot find '{source}'.")

    if source.suffix.lower() in (".xlsx", ".xlsm"):
        header, rows = read_excel(source, sheet)
    elif source.suffix.lower() == ".xls":
        sys.exit(
            f"'{source.name}' is in the old Excel 97-2003 format. Open it in Excel or "
            "LibreOffice and use File > Save As to save it as .xlsx first."
        )
    else:
        header, rows = read_delimited(source)

    columns = unique_names(header)
    width = len(columns)

    # Pad or trim every row so it lines up with the header.
    rows = [(r + [None] * width)[:width] for r in rows]

    forced = {clean_column_name(c, 0) for c in text_columns}
    types = [
        "TEXT" if columns[i] in forced else infer_type([r[i] for r in rows])
        for i in range(width)
    ]

    ddl_columns = ",\n    ".join(f"{c} {t}" for c, t in zip(columns, types))
    ddl = f"CREATE TABLE {table} (\n    {ddl_columns}\n);"

    placeholders = ", ".join("?" for _ in columns)
    insert = f"INSERT INTO {table} ({', '.join(columns)}) VALUES ({placeholders})"

    prepared = [
        tuple(convert(r[i], types[i]) for i in range(width))
        for r in rows
    ]

    conn = sqlite3.connect(database)
    try:
        conn.execute(f"DROP TABLE IF EXISTS {table}")
        conn.execute(ddl)
        conn.executemany(insert, prepared)
        conn.commit()
        count = conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
    finally:
        conn.close()

    if not quiet:
        print(f"\nSource:   {source}" + (f"  (sheet '{sheet}')" if sheet else ""))
        print(f"Database: {database}")
        print(f"\n{ddl}\n")
        print(f"Loaded {count} rows into '{table}'.\n")
    return count


def main():
    parser = argparse.ArgumentParser(
        description="Import a .csv, tab-delimited .txt, or .xlsx file into a SQLite table.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="Example:\n"
               "  python load_data.py data/FrequentFliers.txt frequent_fliers campus_travel.db",
    )
    parser.add_argument("source", help="the data file to read")
    parser.add_argument("table", help="the name to give the table in SQLite")
    parser.add_argument("database", help="the .db file to create or add to")
    parser.add_argument("--sheet", help="worksheet name, for .xlsx files with more than one sheet")
    parser.add_argument("--quiet", action="store_true", help="suppress the summary output")
    parser.add_argument(
        "--text-columns",
        nargs="+",
        default=[],
        metavar="COL",
        help="force these columns to TEXT instead of letting the type be inferred "
             "(use for ZIP codes, ID numbers, and anything else made of digits "
             "that you would never do arithmetic on)",
    )
    args = parser.parse_args()

    load(args.source, args.table, args.database, args.sheet, args.quiet, args.text_columns)


if __name__ == "__main__":
    main()
